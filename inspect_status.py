from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font
import re
import os

# 배관종류 리스트 하드코딩
PIPE_TYPE_LIST = ["입상관", "세대매립관", "세대PD", "세대층상배관", "횡주관"]
# 처리 제외할 배관종류
EXCLUDED_PIPE_TYPES = []  # 입상관도 처리하도록 변경

# 엑셀 파일 로드
wb = load_workbook('sample아이파크_이상배관 보고서_설비사_변준형_0403_working.xlsx')
ws_input = wb['입력창']

# 스타일 정의
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # 하늘색
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 노란색
black_font = Font(color='000000', size=10)  # 검정색, 크기 10

# 1. 모든 배관종류와 범례 추출 (배관종류 -> {배관명 -> 번호})
pipe_types = {}  # 키: 배관종류, 값: {배관명 -> 번호} 사전

# 입력창 7행에서 하드코딩된 배관종류 찾기
col = 1
while col <= ws_input.max_column:
    cell = ws_input.cell(row=7, column=col)
    cell_value = str(cell.value).strip() if cell.value else ""
    
    # 하드코딩된 배관종류 리스트에 있고 제외 리스트에 없는 경우만 처리
    if cell_value in PIPE_TYPE_LIST and cell_value not in EXCLUDED_PIPE_TYPES:
        pipe_type = cell_value
        pipe_start_col = col
        pipe_end_col = None
        
        # 병합셀 범위 확인
        for merged_cell in ws_input.merged_cells.ranges:
            if cell.coordinate in merged_cell:
                pipe_end_col = merged_cell.max_col
                break
        
        # 병합셀 범위가 확인되지 않았다면 다음 열을 확인하여 범위 추정
        if not pipe_end_col:
            next_col = pipe_start_col + 1
            while next_col <= ws_input.max_column:
                next_cell_value = ws_input.cell(row=7, column=next_col).value
                if next_cell_value and next_cell_value != cell.value:
                    pipe_end_col = next_col - 1
                    break
                next_col += 1
            
            # 마지막 열까지 모두 동일한 값이면
            if not pipe_end_col:
                pipe_end_col = ws_input.max_column
        
        # 해당 배관종류의 범례 추출
        legend = {}
        number = 1
        for legend_col in range(pipe_start_col, pipe_end_col + 1):
            val = ws_input.cell(row=8, column=legend_col).value
            if val is not None:
                legend[val] = str(number)
                number += 1
        
        # 배관종류와 범례 저장
        pipe_types[pipe_type] = legend
        
        # 다음 검색을 위해 열 위치 업데이트
        col = pipe_end_col + 1
    else:
        col += 1

# 배관종류 및 범례 정보 출력
print("추출된 배관종류 및 범례:")
for pipe_type, legend in pipe_types.items():
    print(f"{pipe_type}: {legend}")

# 2. 작업폴더에서 동영상 파일 가져오기
video_files = []
work_dir = '.\\작업폴더'
if os.path.exists(work_dir):
    for file in os.listdir(work_dir):
        if file.endswith('.mp4'):
            video_files.append(file)
else:
    print(f"작업폴더({work_dir})가 존재하지 않습니다.")

print(f"발견된 동영상 파일 수: {len(video_files)}")
for i, file in enumerate(video_files[:5], 1):  # 처음 5개만 출력
    print(f"  {i}. {file}")
if len(video_files) > 5:
    print(f"  ... 외 {len(video_files) - 5}개")

# 3. 각 배관종류별로 처리
for pipe_type, pipe_legend in pipe_types.items():
    # 해당 배관종류의 작업현황 시트 찾기 (예: '1.작업현황_세대매립관')
    sheet_name = f"1.작업현황_{pipe_type.strip()}"
    if sheet_name not in wb.sheetnames:
        print(f"시트를 찾을 수 없음: {sheet_name}")
        continue
    
    ws_target = wb[sheet_name]
    print(f"\n처리 중인 시트: {sheet_name}")
    
    # 입상관인 경우 특별 처리
    if pipe_type.strip() == "입상관":
        print("입상관 처리 시작")
        
        # 동 및 라인 열 인덱스 설정
        building_col = 2  # B열
        line_col = 3      # C열
        
        # 3행에서 배관명 추출
        pipe_names = []
        column_to_pipe = {}  # 열 인덱스 -> 배관명 매핑
        
        # 입상관 범례에서 배관명 목록 가져오기
        valid_pipe_names = list(pipe_legend.keys())
        
        # 3행에서 실제 배관명만 추출
        for col_idx in range(1, ws_target.max_column + 1):
            cell_value = ws_target.cell(row=3, column=col_idx).value
            if cell_value and str(cell_value).strip() in valid_pipe_names:
                pipe_names.append(cell_value)
                column_to_pipe[col_idx] = cell_value
        
        # 동영상 파일에서 입상관 정보 추출
        riser_inspections = {}  # 키: (동, 라인), 값: set(배관명)
        
        for fname in video_files:
            # 파일명에서 정보 추출 (예: "102동 1903호 입상관 온수.mp4")
            m = re.match(r"(\d+)동 (\d+)호 (.+?) (\S+)\.mp4", fname)
            if not m:
                continue
            
            building = int(m.group(1))
            unit_str = m.group(2)  # 예: "1903"
            video_pipe_type = m.group(3)  # 배관종류 (예: "입상관")
            pipe_name = m.group(4)  # 배관명 (예: "온수")
            
            # 현재 처리 중인 배관종류와 일치하는지 확인
            if pipe_type.strip() not in video_pipe_type:
                continue
            
            # 호수에서 라인 추출 (마지막 두 자리)
            line = None
            try:
                if len(unit_str) >= 3:
                    line = int(unit_str[-2:])  # 마지막 두 자리
                elif len(unit_str) == 2:
                    line = int(unit_str)       # 두 자리 전체
                elif len(unit_str) == 1:
                    line = int(unit_str)       # 한 자리 전체
                else:
                    continue
            except ValueError:
                continue
            
            key = (building, line)
            if key not in riser_inspections:
                riser_inspections[key] = set()
            
            riser_inspections[key].add(pipe_name)
        
        # 시트에서 행-동-라인 매핑
        all_rows_info = []  # 모든 행의 정보를 저장할 리스트

        # 병합 셀 처리를 위한 이전 동/라인 값 저장
        prev_building_val = None
        prev_line_val = None

        for row_idx in range(5, ws_target.max_row + 1):
            building_val = ws_target.cell(row=row_idx, column=building_col).value
            line_val = ws_target.cell(row=row_idx, column=line_col).value
            
            # 병합 셀 처리: 값이 None이면 이전 값 사용
            if building_val is None:
                building_val = prev_building_val
            else:
                prev_building_val = building_val
            
            if line_val is None:
                line_val = prev_line_val
            else:
                prev_line_val = line_val
            
            if not building_val or not line_val:
                continue
            
            # 문자열로 변환하고 숫자만 추출
            building_str = str(building_val).strip()
            line_str = str(line_val).strip()
            
            # 숫자 추출
            building_match = re.search(r'\d+', building_str)
            line_match = re.search(r'\d+', line_str)
            
            if not building_match or not line_match:
                continue
                
            building = int(building_match.group())
            line = int(line_match.group())
            
            all_rows_info.append((row_idx, building, line))
        
        # 각 동-라인에 대한 검사 정보 처리
        for key, inspected_pipes in riser_inspections.items():
            building, line = key
            
            # 해당 동-라인이 있는 행 찾기
            matching_rows = [row_idx for row_idx, bldg, ln in all_rows_info if bldg == building and ln == line]
            
            if matching_rows:
                for row_idx in matching_rows:
                    # 각 배관명에 대해 처리
                    for col_idx, pipe_name in column_to_pipe.items():
                        if pipe_name in inspected_pipes:
                            cell = ws_target.cell(row=row_idx, column=col_idx)
                            cell.value = "완료"
                            cell.fill = blue_fill
                            cell.font = black_font
                            print(f"[입상관] 동: {building}, 라인: {line}, 배관: {pipe_name} -> 완료 표시")
    else:
        # 기존 처리 (입상관 아닌 경우)
        # 세대별 검사된 배관 번호 모음
        inspected_by_unit = {}  # 키: (동, 호문자열), 값: set(배관번호 문자열들)
        
        for fname in video_files:
            # 파일명에서 정보 추출 (예: "102동 1903호 세대매립관 세탁.mp4")
            m = re.match(r"(\d+)동 (\d+)호 (.+?) (\S+)\.mp4", fname)
            if not m:
                continue
            
            building = int(m.group(1))
            unit_str = m.group(2)  # 예: "1903"
            video_pipe_type = m.group(3)  # 배관종류 (예: "세대매립관")
            pipe_name = m.group(4)  # 배관명 (예: "세탁")
            
            # 현재 처리 중인 배관종류와 일치하는지 확인
            if pipe_type.strip() not in video_pipe_type:
                continue
                
            if (building, unit_str) not in inspected_by_unit:
                inspected_by_unit[(building, unit_str)] = set()
                
            # 범례 사전으로 배관명을 번호로 변환하여 저장
            if pipe_name in pipe_legend:
                inspected_by_unit[(building, unit_str)].add(pipe_legend[pipe_name])
        
        # 세대 위치에 배관번호 기록
        building_info = {
            101: {"start_col": "A", "lines": 4}, 102: {"start_col": "H", "lines": 4},
            103: {"start_col": "O", "lines": 5}, 104: {"start_col": "W", "lines": 6},
            105: {"start_col": "AF", "lines": 4}, 106: {"start_col": "AM", "lines": 4},
            107: {"start_col": "AT", "lines": 4}, 108: {"start_col": "BA", "lines": 4},
            109: {"start_col": "BH", "lines": 4}, 110: {"start_col": "BO", "lines": 4},
        }
        
        for (building, unit_str), pipes in inspected_by_unit.items():
            # 층과 라인 계산
            if len(unit_str) >= 3:
                floor = int(unit_str[:-2])  # 마지막 두 자리를 제외한 부분 (층)
                line = int(unit_str[-2:])   # 마지막 두 자리 (라인)
            else:
                floor = int(unit_str[0])
                line = int(unit_str[1:])  # (3자리 호수 처리)
            
            row = 41 - floor
            start_col_index = column_index_from_string(building_info[building]["start_col"])
            target_col_index = start_col_index + line  # (층 라벨열 + line)
            
            # 해당 열의 마지막행번호-1 행의 셀값 확인
            last_row = ws_target.max_row
            reference_cell = ws_target.cell(row=last_row-1, column=target_col_index)
            reference_value = reference_cell.value
            
            # 셀 값 설정 및 스타일 적용
            cell = ws_target.cell(row=row, column=target_col_index)
            if pipes:
                # 번호들을 "/"로 연결하여 입력
                pipe_str = "/".join(sorted(pipes, key=lambda x: int(x)))
                
                # 참조 셀 값과 비교
                if pipe_str == reference_value:
                    cell.value = "완료"
                    cell.fill = blue_fill
                else:
                    cell.value = pipe_str
                    cell.fill = yellow_fill
                    
                cell.font = black_font
                print(f"[{pipe_type}] 동: {building}, 호수: {unit_str} -> {pipe_str} 처리")
            else:
                cell.value = None  # 점검된 배관 없으면 비워둠

# 수정된 내용을 새 파일로 저장
wb.save('updated_pipe_inspection.xlsx')
print("\n파일 저장 완료: updated_pipe_inspection.xlsx")
