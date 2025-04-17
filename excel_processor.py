#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.drawing.image import Image
import os
import subprocess
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
import re

class ExcelProcessor:
    """
    이미지 파일명을 분석하여 엑셀 파일에 정보를 입력하고 이미지를 삽입하는 클래스
    """
    
    def __init__(self):
        """초기화 함수"""
        self.calculated_width = None
        self.calculated_height = None
        self.processed_images = []
        self.skipped_images = []
        
    def process_images(self, excel_path, image_folder, callback=None):
        """
        이미지 폴더의 이미지를 처리하여 엑셀 파일에 삽입
        
        Args:
            excel_path (str): 엑셀 파일 경로
            image_folder (str): 이미지 폴더 경로
            callback (function): 진행 상황을 알리는 콜백 함수
            
        Returns:
            dict: 처리 결과 (성공 및 실패한 이미지 목록)
        """
        # 파일 존재 여부 확인
        if not os.path.exists(excel_path):
            if callback:
                callback(f"❌ Excel 파일이 존재하지 않습니다: {excel_path}")
            return {"success": False, "error": f"Excel 파일이 존재하지 않습니다: {excel_path}"}

        if not os.path.exists(image_folder):
            if callback:
                callback(f"❌ 이미지 폴더가 존재하지 않습니다: {image_folder}")
            return {"success": False, "error": f"이미지 폴더가 존재하지 않습니다: {image_folder}"}
        
        # 이미지 폴더에서 모든 이미지 파일 가져오기 (jpg, png 모두 처리)
        image_files = [f for f in os.listdir(image_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        if callback:
            callback(f"✓ 처리할 이미지 파일 수: {len(image_files)}")
        
        if not image_files:
            if callback:
                callback("❌ 처리할 이미지 파일이 없습니다.")
            return {"success": False, "error": "처리할 이미지 파일이 없습니다."}
        
        # 엑셀 파일 로드
        wb = None
        try:
            # 엑셀 파일 열기 전에 이미 실행 중인 엑셀 프로세스 확인 및 종료
            self.terminate_excel_processes(excel_path, callback)
            
            # 엑셀 파일 로드
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            if callback:
                callback(f"❌ 엑셀 파일을 열 수 없습니다: {str(e)}")
            return {"success": False, "error": f"엑셀 파일을 열 수 없습니다: {str(e)}"}
        
        # 이미지 처리 결과 저장 변수 초기화
        self.processed_images = []
        self.skipped_images = []
        
        # 처리 결과를 시트별로 저장할 딕셔너리
        results_by_sheet = {}
        
        # 이미지 파일 처리
        for img_file in image_files:
            img_path = os.path.join(image_folder, img_file)
            if os.path.exists(img_path):
                try:
                    # 파일명 분석 (확장자 제외)
                    filename_without_ext = os.path.splitext(img_file)[0]
                    
                    # 파일명 분석: "[동] [호] [배관종류] [배관명]_[이상소견]_[이상위치]"
                    # 먼저 "_"로 주요 부분 분리
                    parts = filename_without_ext.split("_")
                    
                    # _ 기준으로 나눠진 부분이 최소 2개 이상이어야 함
                    if len(parts) >= 2:
                        # 첫 부분은 "[동] [호] [배관종류] [배관명]"
                        location_info = parts[0].strip()
                        
                        # 두 번째 부분은 이상소견
                        issue = parts[1].strip()
                        
                        # 세 번째 부분이 있으면 이상위치, 없으면 공백
                        issue_location = parts[2].strip() if len(parts) > 2 else ""
                        
                        # 공백으로 분리하여 동, 호, 배관종류, 배관명 추출
                        location_parts = location_info.split()
                        
                        # 최소 3개 이상의 부분이 필요 (동, 호, 배관종류)
                        if len(location_parts) >= 3:
                            dong = location_parts[0]  # 첫 번째 부분은 동
                            ho = location_parts[1]    # 두 번째 부분은 호수
                            pipe_type = location_parts[2]  # 세 번째 부분은 배관종류
                            
                            # 네 번째 이후 부분들은 모두 배관명으로 합치기
                            pipe_name = " ".join(location_parts[3:]) if len(location_parts) > 3 else ""
                            
                            # 이상배관LIST 시트 이름 결정 (3.이상배관LIST_[배관종류])
                            list_sheet_name = f"3.이상배관LIST_{pipe_type}"
                            
                            # 이상위치 시트 이름 결정 (2.이상배관위치_[배관종류])
                            location_sheet_name = f"2.이상배관위치_{pipe_type}"
                            
                            # 처리 결과 저장용 시트 키 생성
                            sheet_key = pipe_type
                            if sheet_key not in results_by_sheet:
                                results_by_sheet[sheet_key] = {
                                    "list_sheet": list_sheet_name,
                                    "location_sheet": location_sheet_name,
                                    "processed": [],
                                    "skipped": [],
                                    "location_marked": []
                                }
                            
                            # 이상배관LIST 시트가 있는지 확인
                            if list_sheet_name in wb.sheetnames:
                                list_sheet = wb[list_sheet_name]
                                
                                # 마지막 행 찾기
                                last_row = self.find_last_row(list_sheet)
                                current_row = last_row + 1
                                
                                # 새 행의 높이를 설정 - 기존 템플릿 행의 높이 복사
                                template_row_height = None
                                # 템플릿의 행 높이를 가져오기 (4행을 기준으로 사용)
                                if 4 in list_sheet.row_dimensions:
                                    template_row_height = list_sheet.row_dimensions[4].height
                                    # 새 행에 동일한 높이 적용
                                    list_sheet.row_dimensions[current_row].height = template_row_height
                                
                                # 인덱스(NO) 값 설정 - A열에 순번 입력
                                index_no = current_row - 3  # 헤더를 제외한 인덱스 값 (4행부터 1번)
                                list_sheet[f"A{current_row}"] = index_no
                                
                                # 이미지 생성 일자 가져오기
                                inspection_date = self.get_image_creation_date(img_path)
                                
                                # 데이터 입력
                                list_sheet[f"B{current_row}"] = inspection_date  # 점검일
                                list_sheet[f"C{current_row}"] = dong  # 동명
                                list_sheet[f"D{current_row}"] = ho    # 호수(라인)
                                list_sheet[f"E{current_row}"] = pipe_name  # 배관명
                                list_sheet[f"G{current_row}"] = issue_location  # 이상위치 (G열로 이동)
                                list_sheet[f"H{current_row}"] = issue  # 이상소견 (H열로 이동)
                                
                                # 이미지 삽입 (F열)
                                img = Image(img_path)
                                cell = f"F{current_row}"
                                
                                # 너비를 155픽셀로 고정하고 종횡비 유지
                                width_px = 195  # 너비 155픽셀로 고정
                                # 종횡비 5.16:4.18 유지하여 높이 계산
                                aspect_ratio = 4.18 / 5.16
                                height_px = int(width_px * aspect_ratio)
                                
                                self.calculated_width = width_px
                                self.calculated_height = height_px
                                
                                # 모든 이미지에 동일한 크기 적용
                                img.width = self.calculated_width
                                img.height = self.calculated_height
                                
                                # 이미지 삽입
                                list_sheet.add_image(img, cell)
                                
                                # 간결한 로그 메시지
                                log_msg = f"✓ '{img_file}' -> {list_sheet_name} 시트 {current_row}행"
                                if callback:
                                    callback(log_msg)
                                
                                # 이미지 정보를 저장
                                image_info = {
                                    "file": img_file,
                                    "dong": dong,
                                    "ho": ho,
                                    "pipe_name": pipe_name,
                                    "issue": issue,
                                    "issue_location": issue_location,
                                    "row": current_row
                                }
                                results_by_sheet[sheet_key]["processed"].append(image_info)
                                
                                # 이상위치 시트에 인덱스 추가
                                if location_sheet_name in wb.sheetnames:
                                    # 위치 시트 작업 시작 로그
                                    if callback:
                                        callback(f"\n  === '{dong} {ho}' 이상위치 시트 작업 시작 ===")
                                
                                    # list_sheet_info 명시적으로 구성
                                    sheet_info = {
                                        "sheet_name": list_sheet_name,
                                        "row": current_row,
                                        "d_value": list_sheet[f"D{current_row}"].value,
                                        "pipe_type": pipe_type
                                    }
                                    
                                    # 입상관일 경우 이상위치 정보를 호수 대신 사용
                                    if pipe_type == "입상관":
                                        if callback:
                                            callback(f"  ► 입상관 특별 처리: 이상위치({issue_location})를 호수로 사용")
                                            # D열 값을 직접 출력하여 확인
                                            d_value = list_sheet[f"D{current_row}"].value
                                            callback(f"  ► 이상배관LIST의 D열 값: '{d_value}'")
                                            
                                        # 위치 작업 실행 - 이상위치 정보를 호수로 전달
                                        location_result = self.update_location_sheet(wb[location_sheet_name], dong, issue_location, index_no, callback, sheet_info, wb)
                                    else:
                                        # 다른 배관종류는 기존 방식대로 처리
                                        # 위치 작업 실행 - 원래 호수 정보 사용
                                        location_result = self.update_location_sheet(wb[location_sheet_name], dong, ho, index_no, callback, sheet_info, wb)
                                    
                                    if location_result:
                                        results_by_sheet[sheet_key]["location_marked"].append({
                                            "dong": dong,
                                            "ho": ho if pipe_type != "입상관" else issue_location,
                                            "cell": location_result
                                        })
                                        if callback:
                                            callback(f"  ✓ {location_sheet_name} 시트에 '{dong} {ho if pipe_type != '입상관' else issue_location}' 위치 표시됨 (셀: {location_result})")
                                    else:
                                        if callback:
                                            callback(f"  ❌ {location_sheet_name} 시트에 '{dong} {ho if pipe_type != '입상관' else issue_location}' 위치 표시 실패")
                                    
                                    if callback:
                                        callback(f"  === 이상위치 시트 작업 종료 ===\n")
                                else:
                                    if callback:
                                        callback(f"  ⚠️ {location_sheet_name} 시트가 없습니다")
                                
                                self.processed_images.append(img_file)
                            else:
                                log_msg = f"❌ '{img_file}' -> 파일명 형식 오류 (동,호,배관종류 필요)"
                                if callback:
                                    callback(log_msg)
                                self.skipped_images.append(img_file)
                                results_by_sheet[sheet_key]["skipped"].append({
                                    "file": img_file,
                                    "reason": f"{list_sheet_name} 시트 없음"
                                })
                        else:
                            log_msg = f"❌ '{img_file}' -> 파일명 형식 오류 (최소 1개의 _ 구분자 필요)"
                            if callback:
                                callback(log_msg)
                            self.skipped_images.append(img_file)
                        
                except Exception as e:
                    log_msg = f"❌ '{img_file}' -> 처리 중 오류: {e}"
                    if callback:
                        callback(log_msg)
                    self.skipped_images.append(img_file)
            else:
                log_msg = f"❌ '{img_file}' -> 파일이 존재하지 않음"
                if callback:
                    callback(log_msg)
                self.skipped_images.append(img_file)

        # 처리 결과 출력 (시트별로 구분)
        if callback:
            callback("\n===== 작업 결과 요약 =====")
            callback(f"총 이미지: {len(image_files)}개, 처리 완료: {len(self.processed_images)}개, 처리 실패: {len(self.skipped_images)}개")
            
            if results_by_sheet:
                callback("\n===== 시트별 처리 결과 =====")
                for pipe_type, result in results_by_sheet.items():
                    list_sheet = result["list_sheet"]
                    location_sheet = result["location_sheet"]
                    processed = result["processed"]
                    location_marked = result["location_marked"]
                    
                    callback(f"\n▶ {pipe_type} 배관")
                    callback(f"  • {list_sheet}: {len(processed)}개 이미지 추가")
                    callback(f"  • {location_sheet}: {len(location_marked)}개 위치 표시")
                    
                    if processed:
                        for info in processed:
                            callback(f"    - {info['dong']} {info['ho']} {info['pipe_name']} ({info['issue']}, {info['issue_location']})")

        # 파일 저장
        try:
            # 저장 전에 다시 한번 관련 프로세스 종료
            self.terminate_excel_processes(excel_path, callback)
            
            # 원본 파일에 직접 저장
            wb.save(excel_path)
            
            if callback:
                callback(f"\n✓ 이미지가 삽입된 엑셀 파일이 저장되었습니다: {excel_path}")
            
            # 워크북 객체 닫기 및 메모리 해제
            wb.close()
            wb = None
            
            # 가비지 컬렉션 강제 실행
            import gc
            gc.collect()
            
            # 파일 저장 후 엑셀 프로세스 다시 한번 확인하고 종료
            self.terminate_excel_processes(excel_path, callback)
            
            return {
                "success": True,
                "processed": self.processed_images,
                "skipped": self.skipped_images,
                "total": len(image_files),
                "by_sheet": results_by_sheet
            }
        except Exception as e:
            if callback:
                callback(f"❌ 파일 저장 중 오류 발생: {e}")
                callback("파일이 이미 열려있는 경우 닫은 후 다시 시도하세요.")
            
            # 오류 발생해도 워크북 객체는 닫기
            if wb:
                try:
                    wb.close()
                except:
                    pass
            
            return {
                "success": False,
                "error": f"파일 저장 중 오류 발생: {e}",
                "processed": self.processed_images,
                "skipped": self.skipped_images,
                "total": len(image_files),
                "by_sheet": results_by_sheet
            }
    
    def update_location_sheet(self, sheet, dong, ho, index_no, callback=None, list_sheet_info=None, wb=None):
        """
        이상위치 시트에 표시 추가
        
        Args:
            sheet: 이상위치 시트 객체
            dong: 동 정보 (예: 101동)
            ho: 호수 정보 (예: 205호)
            index_no: 인덱스 번호 (미사용)
            callback: 콜백 함수
            list_sheet_info: 이상배관LIST 시트 정보 (시트명, 현재 행 등)
            wb: 워크북 객체
            
        Returns:
            str or None: 셀 위치 정보 (성공 시) 또는 None (실패 시)
        """
        # 로그 출력 함수 - callback이 없어도 출력하도록 함
        def log_message(message, is_error=False):
            if callback:
                callback(message)
            else:
                # callback이 없는 경우에도 로그 출력
                prefix = "❌ " if is_error else ""
                print(f"{prefix}[이상위치] {message}")
        
        try:
            log_message(f"이상위치 시트 '{sheet.title}'에 '{dong}' '{ho}' 위치 찾는 중...")
            
            # 입상관 특별 처리를 위한 변수
            is_riser_pipe = "입상관" in sheet.title
            is_rooftop = False
            is_basement = False
            use_special_position = False
            
            # 호수에서 숫자만 추출 (예: 205호 -> 205)
            ho_number = ''.join(filter(str.isdigit, ho))
            
            # 입상관 특별 처리 로직
            if is_riser_pipe:
                # 1. 이상위치가 숫자로 시작하는 경우 - 기존 로직대로 처리
                if ho_number:
                    log_message(f"입상관 특별 처리: 이상위치가 숫자로 시작 ('{ho}') - 기존 로직 사용")
                    # ho_number가 이미 추출됨, 기존 로직 계속 사용
                # 2. 이상위치가 '옥상'을 포함하는 경우
                elif '옥상' in ho:
                    log_message(f"입상관 특별 처리: 이상위치에 '옥상' 포함 ('{ho}') - 가장 높은 층 위에 표시")
                    is_rooftop = True
                    use_special_position = True
                # 3. 이상위치가 숫자로 시작하지 않고 '옥상'도 포함하지 않는 경우
                else:
                    log_message(f"입상관 특별 처리: 기타 위치 ('{ho}') - 가장 낮은 층 아래에 표시")
                    is_basement = True
                    use_special_position = True
            
            # 숫자가 없는 경우 (일반 로직) - 에러 발생
            if not ho_number and not use_special_position:
                log_message(f"⚠️ 호수에서 숫자를 추출할 수 없습니다: {ho}", True)
                return None
            
            # 층수와 호수 분리 - 개선된 로직
            # 호수 길이에 따라 층수 처리 (3자리 이상인 경우 앞의 두 자리를 층수로 처리)
            if not use_special_position:
                if len(ho_number) >= 4:  # 1105호와 같은 4자리 이상 호수
                    floor = int(ho_number[:2])  # 앞의 두 자리를 층수로
                    room_number = ho_number[2:]  # 나머지를 호수로
                elif len(ho_number) == 3:  # 105호와 같은 3자리 호수
                    # 3자리인 경우, 첫 번째 자리가 1이고 다음 자리가 0이 아니면 두 자리를 층수로
                    # 예: 105호 -> 10층 5호, 123호 -> 12층 3호
                    if ho_number[0] == '1' and ho_number[1] != '0':
                        floor = int(ho_number[:2])
                        room_number = ho_number[2:]
                    else:
                        floor = int(ho_number[0])
                        room_number = ho_number[1:]
                else:  # 2자리 이하 (25호, 5호 등)
                    floor = int(ho_number[0]) if len(ho_number) >= 2 else 1
                    room_number = ho_number[1:] if len(ho_number) >= 2 else ho_number
                
                log_message(f"호수 분석: {ho} -> {floor}층 {room_number}호")
            
            # 동 이름에서 숫자만 추출 (예: 101동 -> 101)
            dong_number = ''.join(filter(str.isdigit, dong))
            if not dong_number:
                log_message(f"⚠️ 동 이름에서 숫자를 추출할 수 없습니다: {dong}", True)
                return None
            
            log_message(f"동 분석: {dong} -> {dong_number}동")
            
            # 1. 시트 전체의 1층 위치 확인 (A열 기준)
            first_floor_row = None
            for row in range(6, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value  # A열
                if cell_value is not None:
                    # 정확히 "1층"인지 확인 (다른 층과 혼동 방지)
                    cell_str = str(cell_value).strip()
                    if cell_str == "1층":
                        first_floor_row = row
                        log_message(f"시트 전체 1층 위치 찾음: {first_floor_row}행 (A열)")
                        break
            
            if not first_floor_row:
                # B열에서도 확인
                for row in range(6, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=2).value  # B열
                    if cell_value is not None:
                        # 정확히 "1층"인지 확인
                        cell_str = str(cell_value).strip()
                        if cell_str == "1층":
                            first_floor_row = row
                            log_message(f"시트 전체 1층 위치 찾음: {first_floor_row}행 (B열)")
                            break
            
            if not first_floor_row:
                # 1층을 찾지 못한 경우 첫 10개 행의 A, B열 값 출력 (디버깅용)
                log_message("첫 10개 행의 값(디버깅):", False)
                for row in range(1, min(sheet.max_row + 1, 15)):
                    a_val = sheet.cell(row=row, column=1).value
                    b_val = sheet.cell(row=row, column=2).value
                    log_message(f"  행 {row}: A열='{a_val}', B열='{b_val}'")
                
                log_message("⚠️ 이상위치 시트에서 '1층'을 찾을 수 없습니다.", True)
                return None
            
            # 2. 동 찾기 (8행으로 변경) - 이전 4행에서 8행으로 변경됨
            dong_start_col = None
            dong_end_col = None
            
            # 8행의 모든 셀 값 출력 (디버깅용)
            log_message(f"8행 셀 검사 (동:{dong_number} 찾기):", False)
            for col in range(1, min(sheet.max_column + 1, 20)):  # 처음 20개 열만 출력
                cell_value = sheet.cell(row=8, column=col).value
                log_message(f"  열 {col}({get_column_letter(col)}): '{cell_value}'")
            
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=8, column=col).value
                if cell_value is not None and dong_number in str(cell_value):
                    log_message(f"8행 {col}열({get_column_letter(col)})에서 '{dong_number}' 발견: '{cell_value}'")
                    
                    # 병합된 셀 범위 확인
                    merged_cell_found = False
                    for merged_range in sheet.merged_cells.ranges:
                        if merged_range.min_row <= 8 <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                            dong_start_col = merged_range.min_col
                            dong_end_col = merged_range.max_col
                            merged_cell_found = True
                            log_message(f"병합된 셀 범위 발견: {get_column_letter(dong_start_col)}-{get_column_letter(dong_end_col)}")
                            break
                    
                    # 병합된 셀이 없는 경우, 해당 열만 사용
                    if not merged_cell_found:
                        dong_start_col = col
                        dong_end_col = col
                        log_message(f"병합된 셀 없음, 단일 열 사용: {get_column_letter(dong_start_col)}")
                    
                    log_message(f"{dong} 동 위치 찾음: {get_column_letter(dong_start_col)}-{get_column_letter(dong_end_col)}")
                    break
            
            if not dong_start_col:
                log_message(f"⚠️ 이상위치 시트에서 '{dong}'을(를) 찾을 수 없습니다.", True)
                return None
            
            # 3. 층수에 따른 행 계산 (시트 전체 1층 위치 기준)
            # 일반 케이스 또는 입상관 특별 처리에 따라 다르게 계산
            if use_special_position:
                # 옥상인 경우 가장 높은 층을 찾아 그 위 행을 사용
                if is_rooftop:
                    # A열에서 "층"이 포함된 모든 행을 찾아 가장 높은 층 찾기
                    highest_floor = 1
                    highest_floor_row = first_floor_row
                    
                    # A열과 B열에서 층 정보 검색
                    for row in range(6, first_floor_row + 20):  # 1층 기준으로 위쪽 20행까지만 검색
                        for col in [1, 2]:  # A열, B열 검사
                            cell_value = sheet.cell(row=row, column=col).value
                            if cell_value is not None and "층" in str(cell_value):
                                # 숫자 추출
                                floor_match = re.search(r'(\d+)층', str(cell_value))
                                if floor_match:
                                    floor_num = int(floor_match.group(1))
                                    if floor_num > highest_floor:
                                        highest_floor = floor_num
                                        highest_floor_row = row
                    
                    # 가장 높은 층 바로 위 행 사용
                    floor_row = highest_floor_row - 1
                    log_message(f"옥상 처리: 가장 높은 층({highest_floor}층)의 위치는 {highest_floor_row}행, 옥상 위치로 {floor_row}행 사용")
                
                # 지하인 경우 가장 낮은 층을 찾아 그 아래 행을 사용
                elif is_basement:
                    # 1층 바로 아래 행 사용
                    floor_row = first_floor_row + 1
                    log_message(f"지하 처리: 1층 위치는 {first_floor_row}행, 지하 위치로 {floor_row}행 사용")
            else:
                # 일반 케이스: 층수에 따른 행 계산
                floor_row = first_floor_row - (floor - 1)
            
            if floor_row < 6 or floor_row > sheet.max_row:  # 행 범위 확인
                log_message(f"⚠️ 계산된 행 위치({floor_row})가 유효하지 않습니다.", True)
                return None
            
            # 층수 계산 결과 로그
            if use_special_position:
                log_message(f"특별 위치 행 계산: {floor_row}행 ('{ho}')")
            else:
                log_message(f"{floor}층 행 위치 계산: {floor_row}행")
            
            # 4. 호수 찾기 (각 동 열 범위 내에서)
            room_col = None
            
            # 입상관 특별 처리의 경우 D열(4열)을 기준으로 라인 찾기
            if use_special_position:
                # 호수(라인) 찾기 - D열 참조
                try:
                    # 디버깅: 원본 호수와 입력받은 정보 확인
                    log_message(f"입상관 처리 시작: 원본 호수='{ho}', D열 정보 확인 중...")
                    
                    # D열 값 초기화
                    d_val = None
                    line_number = None
                    
                    # 이상배관LIST 시트에서 D열 값을 가져옴
                    if list_sheet_info and wb:
                        log_message(f"list_sheet_info 내용: {list_sheet_info}")
                        
                        # 옵션 1: 미리 저장된 D열 값 사용
                        if "d_value" in list_sheet_info and list_sheet_info["d_value"] is not None:
                            d_val = list_sheet_info["d_value"]
                            log_message(f"이상배관LIST의 D열 값(미리 저장됨): '{d_val}'")
                        
                        # 옵션 2: 직접 시트에서 가져오기
                        elif "sheet_name" in list_sheet_info and "row" in list_sheet_info:
                            sheet_name = list_sheet_info["sheet_name"]
                            row_num = list_sheet_info["row"]
                            
                            if sheet_name in wb.sheetnames:
                                target_sheet = wb[sheet_name]
                                cell_value = target_sheet.cell(row=row_num, column=4).value  # D열
                                d_val = cell_value
                                log_message(f"이상배관LIST 시트({sheet_name})의 {row_num}행 D열 값: '{d_val}'")
                            else:
                                log_message(f"⚠️ 시트를 찾을 수 없음: {sheet_name}")
                    
                    # D열 값이 없는 경우 처리
                    if d_val is None:
                        log_message(f"⚠️ D열 값을 가져올 수 없음, 기본값 사용")
                        room_col = dong_start_col  # 기본값으로 첫 열 사용
                    else:
                        # D열 값이 있는 경우 처리
                        log_message(f"D열 값 추출 시작: '{d_val}'")
                        
                        # 문자열로 변환
                        d_val_str = str(d_val)
                        
                        # "호" 문자 제거 및 공백 제거
                        d_val_clean = d_val_str.replace("호", "").strip()
                        
                        # 숫자만 추출
                        d_number = ''.join(filter(str.isdigit, d_val_clean))
                        
                        log_message(f"D열 값 전처리: 원본='{d_val}', 정리='{d_val_clean}', 숫자만='{d_number}'")
                        
                        # 숫자가 있는 경우만 처리
                        if d_number:
                            try:
                                # 단순히 숫자를 정수로 변환
                                line_number = int(d_number)
                                log_message(f"✓ D열 라인값 추출 성공: {d_number} → 라인 {line_number}")
                            except ValueError:
                                log_message(f"⚠️ D열 숫자 변환 실패: '{d_number}'")
                                # 값 추출 실패 시 첫 번째 열 사용
                                room_col = dong_start_col
                            
                            # 라인 번호가 추출된 경우
                            if line_number is not None:
                                # 라인 번호에 해당하는 열 계산
                                column_offset = line_number - 1  # 1-기반 인덱스를 0-기반으로 변환
                                
                                # 동 범위 내에서 해당 호수 위치 계산
                                if 0 <= column_offset <= (dong_end_col - dong_start_col):
                                    # 계산된 위치가 동 범위 내에 있는 경우
                                    room_col = dong_start_col + column_offset
                                    log_message(f"라인 {line_number}의 열 위치 계산: {get_column_letter(room_col)}열")
                                else:
                                    # 범위를 벗어난 경우
                                    if line_number <= 0:
                                        room_col = dong_start_col  # 첫 열 사용
                                        log_message(f"라인 {line_number}이 최소값 이하, 첫 번째 열 사용: {get_column_letter(room_col)}")
                                    elif line_number > (dong_end_col - dong_start_col + 1):
                                        room_col = dong_end_col  # 마지막 열 사용
                                        log_message(f"라인 {line_number}이 범위 초과, 마지막 열 사용: {get_column_letter(room_col)}")
                        else:
                            # 숫자가 없는 경우
                            log_message(f"⚠️ D열 값에서 숫자를 추출할 수 없음: '{d_val}'")
                            room_col = dong_start_col  # 기본값으로 첫 열 사용
                
                except Exception as e:
                    import traceback
                    error_detail = traceback.format_exc()
                    log_message(f"⚠️ 라인 정보 추출 중 오류: {str(e)}", True)
                    log_message(f"상세 오류 정보: {error_detail}", True)
                    # 오류 발생 시 기본값으로 첫 번째 열 사용
                    room_col = dong_start_col
                    log_message(f"오류 발생, 첫 번째 열({get_column_letter(room_col)}) 사용")
            else:
                # 호수가 0으로 시작하면 0을 제거한 형태로 변환
                room_number_clean = room_number.lstrip('0')
                if room_number_clean == '': room_number_clean = '0'  # 0만 있는 경우 처리
                
                try:
                    # 호수를 정수로 변환 (예: "05" -> 5, "12" -> 12)
                    room_number_int = int(room_number_clean)
                    
                    # 동 범위 내에서 호수 위치 계산 방식 적용
                    # 1호는 동 범위의 첫 번째 열, 2호는 두 번째 열... 식으로 계산
                    column_offset = room_number_int - 1  # 호수에서 1을 빼서 0부터 시작하는 인덱스로 변환
                    
                    log_message(f"호수 계산: {room_number} -> {room_number_int}호 (오프셋: {column_offset})")
                    log_message(f"동 범위: {dong_start_col}-{dong_end_col} ({get_column_letter(dong_start_col)}-{get_column_letter(dong_end_col)}), 총 {dong_end_col - dong_start_col + 1}개 열")
                    
                    # 동 범위 내에서 해당 호수 위치 계산
                    if 0 <= column_offset <= (dong_end_col - dong_start_col):
                        # 계산된 위치가 동 범위 내에 있는 경우
                        room_col = dong_start_col + column_offset
                        log_message(f"{room_number_clean}호 열 위치 계산: {get_column_letter(room_col)}열 (범위 내)")
                    else:
                        # 계산된 위치가 범위를 벗어나는 경우 경고
                        log_message(f"⚠️ 계산된 {room_number_clean}호 위치가 동 범위를 벗어납니다 ({column_offset+1}번째 위치, 범위는 1-{dong_end_col-dong_start_col+1})", True)
                        
                        # 범위를 벗어난 경우 처리 방법:
                        if room_number_int <= 0:
                            # 0이하 호수는 첫 번째 열 사용
                            room_col = dong_start_col
                            log_message(f"범위를 벗어나므로 첫 번째 열({get_column_letter(room_col)})로 조정")
                        elif room_number_int > (dong_end_col - dong_start_col + 1):
                            # 범위를 초과하는 호수는 마지막 열 사용
                            room_col = dong_end_col
                            log_message(f"범위를 벗어나므로 마지막 열({get_column_letter(room_col)})로 조정")
                
                except ValueError:
                    # 호수를 숫자로 변환할 수 없는 경우 (예: 비표준 호수)
                    log_message(f"⚠️ 호수 '{room_number}'를 숫자로 해석할 수 없습니다.", True)
                    return None
            
            # 호수를 찾지 못한 경우
            if not room_col:
                if use_special_position:
                    log_message(f"⚠️ 특별 위치 처리 중 열 위치를 찾을 수 없습니다.", True)
                else:
                    log_message(f"⚠️ 이상위치 시트에서 '{room_number}호'를 찾을 수 없습니다.", True)
                return None
            
            # 5. 해당 셀 배경색을 노란색으로 변경하고 인덱스 값 입력
            target_cell = sheet.cell(row=floor_row, column=room_col)
            
            # 셀의 현재 배경 확인 - 디버깅 로그 추가
            current_fill = target_cell.fill
            is_already_yellow = False
            
            # 노란색 배경인지 확인 (RGB 값과 fill_type 확인)
            log_message(f"현재 셀 배경색 속성: {current_fill}")
            
            # 배경색 확인 방법 개선
            try:
                if hasattr(current_fill, 'start_color') and hasattr(current_fill.start_color, 'rgb'):
                    rgb_value = current_fill.start_color.rgb
                    log_message(f"현재 셀 RGB 값: {rgb_value}")
                    
                    # RGB 값이 없거나 null이면 배경색이 없는 것으로 간주
                    if not rgb_value or rgb_value == '00000000':
                        is_already_yellow = False
                    else:
                        # FFFF00이 포함된 모든 형식 인식 (접두사 무시)
                        is_already_yellow = 'FFFF00' in rgb_value
                        
                    log_message(f"노란색 배경 여부: {is_already_yellow}")
                elif hasattr(current_fill, 'fgColor') and hasattr(current_fill.fgColor, 'rgb'):
                    # 이전 버전 openpyxl에서는 fgColor를 사용할 수 있음
                    rgb_value = current_fill.fgColor.rgb
                    log_message(f"현재 셀 fgColor RGB 값: {rgb_value}")
                    
                    # RGB 값이 없거나 null이면 배경색이 없는 것으로 간주
                    if not rgb_value or rgb_value == '00000000':
                        is_already_yellow = False
                    else:
                        # FFFF00이 포함된 모든 형식 인식 (접두사 무시)
                        is_already_yellow = 'FFFF00' in rgb_value
                        
                    log_message(f"노란색 배경 여부: {is_already_yellow}")
                else:
                    log_message("셀 배경색 속성을 확인할 수 없습니다.")
            except Exception as e:
                log_message(f"배경색 확인 중 오류 발생: {str(e)}")
            
            # 현재 셀 값 확인
            log_message(f"현재 셀 값: '{target_cell.value}'")
            
            # 노란색 배경 설정
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            target_cell.fill = yellow_fill
            
            # 이상배관LIST의 NO 컬럼 값(index_no)을 셀에 입력
            from openpyxl.styles import Font
            
            # 현재 셀 값
            current_value = target_cell.value
            
            # 디버깅을 위한 배경색 확인 로그 추가
            log_message(f"최종 배경색 확인 결과: {'노란색임' if is_already_yellow else '노란색 아님'}")
            
            # 이미 노란색 배경이었으면 인덱스 목록으로 간주
            if is_already_yellow and current_value is not None:
                current_value_str = str(current_value)
                # 쉼표로 구분된 값들 추출
                parts = [part.strip() for part in current_value_str.split(',')]
                
                # 중복 체크 - 같은 인덱스가 이미 있는지 확인
                if str(index_no) not in parts:
                    # 없으면 추가 (쉼표로 구분)
                    target_cell.value = f"{current_value_str},{index_no}"
                    log_message(f"기존 인덱스 목록 '{current_value_str}'에 인덱스 {index_no} 추가")
                else:
                    log_message(f"인덱스 {index_no}가 이미 존재함, 중복 추가하지 않음")
            else:
                # 노란색 배경이 아니었으면 호수 정보로 간주하고 인덱스로 대체
                target_cell.value = str(index_no)
                if current_value is not None:
                    log_message(f"호수 정보 '{current_value}'를 인덱스 {index_no}로 대체")
                else:
                    log_message(f"빈 셀에 인덱스 {index_no} 입력")
            
            # 텍스트 색상을 검정색으로 설정하고 크기를 10으로 지정, 가운데 정렬 추가
            target_cell.font = Font(color="000000", size=10)
            target_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 셀 위치 정보 (예: A1)
            cell_address = f"{get_column_letter(room_col)}{floor_row}"
            
            # 특별 처리 케이스와 일반 케이스별로 다른 로그 메시지 출력
            if use_special_position:
                if is_rooftop:
                    position_desc = "옥상"
                elif is_basement:
                    position_desc = "최하단"
                else:
                    position_desc = "특별 위치"
                
                log_message(f"✓ 이상위치 시트에 '{dong}' {position_desc} 위치 셀({cell_address})의 배경색을 노란색으로 변경하고 인덱스 {index_no}을(를) 입력했습니다.")
            else:
                log_message(f"✓ 이상위치 시트에 '{dong}' {floor}층 {room_number}호 셀({cell_address})의 배경색을 노란색으로 변경하고 인덱스 {index_no}을(를) 입력했습니다.")
            
            # 셀 위치 정보 반환
            return cell_address
                
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            log_message(f"❌ 이상위치 시트 업데이트 중 오류: {str(e)}", True)
            log_message(f"상세 오류 정보: {error_detail}", True)
            return None
    
    def get_cell_dimensions(self, sheet, col_letter, row_number):
        """셀의 크기(픽셀)를 계산"""
        # 열 너비 가져오기 (포인트 단위)
        col_width = sheet.column_dimensions[col_letter].width
        
        # 행 높이 가져오기 (포인트 단위)
        row_height = sheet.row_dimensions[row_number].height
        
        # 기본값이 설정되지 않은 경우 기본값 사용
        if col_width is None:
            col_width = 8.43  # Excel의 기본 열 너비
        if row_height is None:
            row_height = 15.0  # Excel의 기본 행 높이
        
        # 포인트를 픽셀로 변환 (대략적인 변환)
        col_width_px = col_width * 9  # 조정된 변환 계수
        row_height_px = row_height * 1.5  # 변환 계수 조정
        
        return col_width_px, row_height_px

    def get_image_creation_date(self, img_path):
        """이미지 파일의 수정일자를 MM/DD 형식으로 반환"""
        if os.path.exists(img_path):
            # timestamp = os.path.getctime(img_path)  # 파일 생성 시간 가져오기
            timestamp = os.path.getmtime(img_path)  # 파일 수정 시간 가져오기
            date_obj = datetime.fromtimestamp(timestamp)
            return date_obj.strftime("%m/%d")  # MM/DD 형식으로 반환
        return ""

    def find_last_row(self, sheet, date_col='B'):
        """시트에서 점검일 컬럼(B)에 값이 있는 마지막 행 번호 반환"""
        last_row = 3  # 3행까지는 헤더, 4행부터 데이터 시작
        
        for row in range(4, sheet.max_row + 1):  # 4행부터 데이터 확인
            cell_value = sheet[f"{date_col}{row}"].value
            if cell_value is not None:
                last_row = row
        
        return last_row
    
    def open_excel_file(self, excel_path, callback=None):
        """엑셀 파일 열기"""
        try:
            # 엑셀 실행 명령
            excel_app = r'C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE'
            if not os.path.exists(excel_app):
                excel_app = r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE'
            if not os.path.exists(excel_app):
                # 기본 연결 프로그램으로 열기
                os.startfile(excel_path)
            else:
                # 엑셀로 직접 열기
                subprocess.Popen([excel_app, excel_path])
            
            if callback:
                callback(f"✓ 엑셀 파일이 열렸습니다. 처리된 내용을 확인하세요.")
            return True
        except Exception as e:
            if callback:
                callback(f"❌ 엑셀 파일을 여는 중 오류가 발생했습니다: {e}")
                callback(f"수동으로 파일을 열고 확인하세요: {excel_path}")
            return False

    def terminate_excel_processes(self, excel_path, callback=None):
        """
        엑셀 프로세스를 안전하게 종료
        
        Args:
            excel_path (str): 엑셀 파일 경로 (로그용)
            callback (function): 콜백 함수
            
        Returns:
            bool: 성공 여부
        """
        try:
            import subprocess
            import time
            import sys
            
            # 파일 정보 (로그용)
            file_name = os.path.basename(excel_path)
            
            if callback:
                callback(f"→ 엑셀 프로세스 확인 중...")
            
            # 먼저 특정 프로세스가 파일을 열고 있는지 확인 (Windows 전용)
            if sys.platform == 'win32':
                try:
                    # 1. 모든 엑셀 프로세스 확인
                    process = subprocess.Popen(
                        ['tasklist', '/FI', 'IMAGENAME eq EXCEL.EXE', '/FO', 'CSV'], 
                        stdout=subprocess.PIPE, 
                        stderr=subprocess.PIPE,
                        shell=True
                    )
                    output, _ = process.communicate()
                    output_str = output.decode('utf-8', errors='ignore')
                    
                    # Excel 프로세스 있는지 확인
                    if 'EXCEL.EXE' in output_str:
                        if callback:
                            callback(f"→ Excel 프로세스 발견, 종료 시도 중...")
                        
                        # Excel 프로세스 모두 종료
                        subprocess.run(
                            ['taskkill', '/IM', 'EXCEL.EXE', '/F'], 
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE,
                            shell=True
                        )
                        
                        # 프로세스가 완전히 종료될 때까지 짧은 대기
                        time.sleep(1.5)
                        
                        if callback:
                            callback(f"✓ Excel 프로세스를 종료했습니다")
                    else:
                        if callback:
                            callback(f"✓ 실행 중인 Excel 프로세스가 없습니다")
                except Exception as e:
                    if callback:
                        callback(f"⚠️ Excel 프로세스 확인/종료 중 오류: {str(e)}")
                
                # COM 객체 정리 (추가 안전장치)
                try:
                    import pythoncom
                    pythoncom.CoUninitialize()
                except:
                    pass
            else:
                # Windows가 아닌 경우
                if callback:
                    callback(f"✓ Windows 환경이 아닙니다. 프로세스 종료를 건너뜁니다.")
            
            return True
            
        except Exception as e:
            if callback:
                callback(f"⚠️ Excel 프로세스 종료 중 오류: {str(e)}")
            return False 

    def update_work_status(self, excel_path, video_folder, callback=None):
        """
        동영상 파일명을 기준으로 작업 현황 엑셀 시트 업데이트
        
        Args:
            excel_path (str): 엑셀 파일 경로
            video_folder (str): 동영상 파일이 있는 폴더 경로
            callback (function): 진행 상황을 알리는 콜백 함수
            
        Returns:
            dict: 처리 결과 (성공 및 실패한 파일 목록)
        """
        # 파일 존재 여부 확인
        if not os.path.exists(excel_path):
            if callback:
                callback(f"❌ Excel 파일이 존재하지 않습니다: {excel_path}")
            return {"success": False, "error": f"Excel 파일이 존재하지 않습니다: {excel_path}"}

        if not os.path.exists(video_folder):
            if callback:
                callback(f"❌ 동영상 폴더가 존재하지 않습니다: {video_folder}")
            return {"success": False, "error": f"동영상 폴더가 존재하지 않습니다: {video_folder}"}
        
        # 동영상 폴더에서 모든 동영상 파일 가져오기
        video_files = [f for f in os.listdir(video_folder) 
                      if f.lower().endswith(('.mp4', '.avi', '.mov', '.wmv', '.mkv', '.flv'))]
        
        if callback:
            callback(f"✓ 처리할 동영상 파일 수: {len(video_files)}")
        
        if not video_files:
            if callback:
                callback("❌ 처리할 동영상 파일이 없습니다.")
            return {"success": False, "error": "처리할 동영상 파일이 없습니다."}
        
        # 엑셀 파일 로드
        wb = None
        try:
            # 엑셀 파일 열기 전에 이미 실행 중인 엑셀 프로세스 확인 및 종료
            self.terminate_excel_processes(excel_path, callback)
            
            # 엑셀 파일 로드
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            if callback:
                callback(f"❌ 엑셀 파일을 열 수 없습니다: {str(e)}")
            return {"success": False, "error": f"엑셀 파일을 열 수 없습니다: {str(e)}"}
        
        # 처리 결과 저장 변수 초기화
        processed_files = []
        skipped_files = []
        
        # 작업 현황 시트 목록 확인 (1.작업현황_ 으로 시작하는 모든 시트)
        status_sheets = [sheet_name for sheet_name in wb.sheetnames if sheet_name.startswith("1.작업현황_")]
        
        if not status_sheets:
            if callback:
                callback("❌ 작업 현황 시트(1.작업현황_로 시작하는 시트)가 없습니다.")
            return {"success": False, "error": "작업 현황 시트가 없습니다."}
        
        # 작업 현황 결과를 시트별로 저장할 딕셔너리
        results_by_sheet = {}
        
        # 작업 현황 시트 정보 출력
        if callback:
            callback(f"✓ 작업 현황 시트 목록:")
            for i, sheet_name in enumerate(status_sheets, 1):
                callback(f"  {i}. {sheet_name}")
            callback("")
        
        # 현재 날짜 가져오기 (배관검사 완료일 입력용)
        today = datetime.now().strftime("%Y-%m-%d")
        
        # 완료 상태 표시용 배경색 설정 (연한 녹색)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # 동영상 파일 처리
        for video_file in video_files:
            if callback:
                callback(f"\n=== '{video_file}' 처리 중... ===")
            
            try:
                # 파일명 분석 (확장자 제외)
                filename_without_ext = os.path.splitext(video_file)[0]
                
                # 파일명 분석: "[동] [호] [배관종류] [배관명]"
                parts = filename_without_ext.split()
                
                # 최소 3개 이상의 부분이 필요 (동, 호, 배관종류)
                if len(parts) >= 3:
                    dong = parts[0]  # 첫 번째 부분은 동
                    ho = parts[1]    # 두 번째 부분은 호수
                    pipe_type = parts[2]  # 세 번째 부분은 배관종류
                    
                    # 네 번째 이후 부분들은 모두 배관명으로 합치기
                    pipe_name = " ".join(parts[3:]) if len(parts) > 3 else ""
                    
                    # 작업 현황 시트 찾기 (1.작업현황_[배관종류])
                    status_sheet_name = f"1.작업현황_{pipe_type}"
                    
                    # 결과 저장용 시트 키 생성
                    sheet_key = pipe_type
                    if sheet_key not in results_by_sheet:
                        results_by_sheet[sheet_key] = {
                            "status_sheet": status_sheet_name,
                            "processed": [],
                            "skipped": []
                        }
                    
                    # 해당 배관종류에 맞는 작업 현황 시트가 있는지 확인
                    if status_sheet_name in wb.sheetnames:
                        status_sheet = wb[status_sheet_name]
                        
                        # 동, 호 정보로 행을 찾기
                        found = False
                        row_to_update = None
                        
                        # 동호수에서 숫자만 추출
                        dong_number = ''.join(filter(str.isdigit, dong))
                        ho_number = ''.join(filter(str.isdigit, ho))
                        
                        if callback:
                            callback(f"  검색할 동호수 정보: 동={dong}(숫자:{dong_number}), 호={ho}(숫자:{ho_number})")
                        
                        # 시트 내용을 탐색하여 동, 호가 맞는 행 찾기
                        for row in range(4, status_sheet.max_row + 1):  # 4행부터 시작 (헤더 제외)
                            # C열은 동, D열은 호
                            cell_dong = status_sheet.cell(row=row, column=3).value  # C열
                            cell_ho = status_sheet.cell(row=row, column=4).value  # D열
                            
                            # 셀 값이 None인 경우 건너뛰기
                            if cell_dong is None or cell_ho is None:
                                continue
                                
                            # 셀에서 숫자만 추출
                            cell_dong_number = ''.join(filter(str.isdigit, str(cell_dong)))
                            cell_ho_number = ''.join(filter(str.isdigit, str(cell_ho)))
                            
                            # 디버깅 정보 (처음 5개 행만)
                            if row < 9:
                                if callback:
                                    callback(f"  행 {row}: 동={cell_dong}(숫자:{cell_dong_number}), 호={cell_ho}(숫자:{cell_ho_number})")
                            
                            # 숫자 부분만 비교하여 일치 여부 확인
                            if cell_dong_number == dong_number and cell_ho_number == ho_number:
                                row_to_update = row
                                found = True
                                if callback:
                                    callback(f"  ✓ {row}행에서 일치하는 정보 발견: {cell_dong} / {cell_ho}")
                                break
                        
                        if found and row_to_update:
                            # H열 (배관검사)에 완료 표시
                            status_sheet.cell(row=row_to_update, column=8).value = "완료"  # H열
                            status_sheet.cell(row=row_to_update, column=8).fill = green_fill
                            
                            # I열 (배관검사 완료일)에 날짜 입력
                            status_sheet.cell(row=row_to_update, column=9).value = today  # I열
                            
                            if callback:
                                callback(f"✓ '{dong} {ho}' 정보 업데이트 완료 ({status_sheet_name} 시트 {row_to_update}행)")
                            
                            # 처리 결과 저장
                            processed_files.append(video_file)
                            results_by_sheet[sheet_key]["processed"].append({
                                "file": video_file,
                                "dong": dong,
                                "ho": ho,
                                "row": row_to_update
                            })
                        else:
                            if callback:
                                callback(f"❌ '{dong} {ho}' 정보를 찾을 수 없습니다")
                            
                            # 실패 결과 저장
                            skipped_files.append(video_file)
                            results_by_sheet[sheet_key]["skipped"].append({
                                "file": video_file,
                                "reason": f"'{dong} {ho}' 정보를 시트에서 찾을 수 없음"
                            })
                    else:
                        if callback:
                            callback(f"❌ '{status_sheet_name}' 시트가 없습니다")
                        
                        # 실패 결과 저장
                        skipped_files.append(video_file)
                        results_by_sheet[sheet_key]["skipped"].append({
                            "file": video_file,
                            "reason": f"'{status_sheet_name}' 시트 없음"
                        })
                else:
                    if callback:
                        callback(f"❌ '{video_file}' 파일명 형식 오류 (최소 동, 호, 배관종류 정보 필요)")
                    
                    # 실패 결과 저장
                    skipped_files.append(video_file)
                    if callback:
                        callback(f"  파일명 형식은 '[동] [호] [배관종류] [배관명]' 이어야 합니다")
                    
            except Exception as e:
                if callback:
                    callback(f"❌ '{video_file}' 처리 중 오류: {str(e)}")
                
                # 실패 결과 저장
                skipped_files.append(video_file)
        
        # 처리 결과 출력 (시트별로 구분)
        if callback:
            callback("\n===== 작업 결과 요약 =====")
            
            for sheet_key, results in results_by_sheet.items():
                total_processed = len(results["processed"])
                total_skipped = len(results["skipped"])
                
                callback(f"\n- {sheet_key} 배관:")
                callback(f"  • 작업 현황 시트: {results['status_sheet']}")
                callback(f"  • 업데이트 성공: {total_processed}개")
                callback(f"  • 업데이트 실패: {total_skipped}개")
        
        # 변경사항 저장
        try:
            wb.save(excel_path)
            if callback:
                callback(f"✓ 엑셀 파일 저장 완료: {os.path.basename(excel_path)}")
        except Exception as e:
            if callback:
                callback(f"❌ 엑셀 파일 저장 중 오류: {str(e)}")
            return {"success": False, "error": f"엑셀 파일 저장 중 오류: {str(e)}"}
        
        # 처리 결과 반환
        return {
            "success": True,
            "total": len(video_files),
            "processed": processed_files,
            "skipped": skipped_files,
            "details": results_by_sheet
        } 